from __future__ import annotations

import csv
import json
import re
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from config import get_analytics_glossary_json_path, get_analytics_glossary_xlsx_path
except ImportError:  # pragma: no cover
    from .config import get_analytics_glossary_json_path, get_analytics_glossary_xlsx_path


REPO_ROOT = Path(__file__).resolve().parent
DATA_DIR = REPO_ROOT / "DATA"
DEFAULT_GLOSSARY_WORKBOOK = DATA_DIR / "RISK_MultiScenario_Unified_Business_Glossary 1.xlsx"
DEFAULT_GLOSSARY_JSON = DATA_DIR / "Business_Glossary_Output.json"
GLOSSARY_SHEET_CANDIDATES = ("Glossary_All", "Glosarry_all")
REQUIREMENT_SHEET_CANDIDATES = ("Requirement",)

ABBREVIATION_MAP = {
    "acct": "account",
    "addr": "address",
    "amt": "amount",
    "arr": "arrangement",
    "bal": "balance",
    "cd": "code",
    "cnt": "count",
    "coll": "collateral",
    "cust": "customer",
    "desc": "description",
    "dt": "date",
    "flg": "flag",
    "identification": "identifier",
    "id": "identifier",
    "ind": "indicator",
    "int": "interest",
    "kyc": "know_your_customer",
    "ln": "loan",
    "nbr": "number",
    "nm": "name",
    "no": "number",
    "num": "number",
    "pct": "percent",
    "prov": "provision",
    "qty": "quantity",
    "ref": "reference",
    "seg": "segment",
    "stat": "status",
    "sts": "status",
    "txn": "transaction",
    "typ": "type",
}

REQUIREMENT_REQUIRED_COLUMNS = {
    "business_attribute": "Business Attribute",
    "business_definition": "Business Definition",
}
REQUIREMENT_OPTIONAL_COLUMNS = {
    "region": "Region",
    "value_stream": "Value Stream",
    "brief_problem_statement": "Brief Problem Statement / Use Case",
    "system_requirements": "System Requirements",
}


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value != value:
        return ""
    return str(value).strip()


def _safe_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    text = _safe_text(value).lower()
    if text in {"true", "yes", "y", "1"}:
        return True
    if text in {"false", "no", "n", "0"}:
        return False
    return None


def _normalize_name(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", _safe_text(text).lower()).strip("_")


def _display_text(text: str) -> str:
    return _normalize_name(text).replace("_", " ")


def _expand_attribute_name(text: str) -> str:
    expanded_tokens: list[str] = []
    for token in _normalize_name(text).split("_"):
        if not token:
            continue
        replacement = ABBREVIATION_MAP.get(token, token)
        expanded_tokens.extend(part for part in replacement.split("_") if part)
    return "_".join(expanded_tokens)


def _extract_lineage_assets(text: str) -> list[str]:
    assets: list[str] = []
    for token in re.findall(r"\b[A-Z][A-Z0-9_]{2,}\b", _safe_text(text)):
        if token in {"MDA", "GDA", "CDA", "UK"} or token in assets:
            continue
        assets.append(token)
    return assets


def _is_join_key(attribute_name: str) -> bool:
    key = _normalize_name(attribute_name)
    return key.endswith("_id") or key.endswith("_key") or key.endswith("_ref_id")


def _resolve_glossary_workbook_path() -> Path:
    configured = _safe_text(get_analytics_glossary_xlsx_path())
    if configured:
        path = Path(configured)
        if path.exists():
            return path

    if DEFAULT_GLOSSARY_WORKBOOK.exists():
        return DEFAULT_GLOSSARY_WORKBOOK

    raise FileNotFoundError(
        "Analytics glossary workbook not found. Set ANALYTICS_GLOSSARY_XLSX_PATH or place "
        f"{DEFAULT_GLOSSARY_WORKBOOK.name} inside the DATA folder."
    )


def _resolve_glossary_json_output_path() -> Path:
    configured = _safe_text(get_analytics_glossary_json_path())
    if configured:
        return Path(configured)
    return DEFAULT_GLOSSARY_JSON


def _resolve_sheet_name(sheet_names: list[str], candidates: tuple[str, ...]) -> str:
    normalized_map = {_normalize_name(name): name for name in sheet_names}
    for candidate in candidates:
        match = normalized_map.get(_normalize_name(candidate))
        if match:
            return match
    raise ValueError(
        f"Required worksheet not found. Expected one of: {', '.join(candidates)}. "
        f"Available sheets: {', '.join(sheet_names)}"
    )


def _open_workbook_rows_from_path(path: Path, sheet_candidates: tuple[str, ...]) -> tuple[list[str], list[tuple[Any, ...]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = _resolve_sheet_name(list(workbook.sheetnames), sheet_candidates)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [_safe_text(value) for value in rows[0]]
    return headers, rows[1:]


def _open_workbook_rows_from_bytes(file_bytes: bytes, sheet_candidates: tuple[str, ...]) -> tuple[list[str], list[tuple[Any, ...]]]:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_name = _resolve_sheet_name(list(workbook.sheetnames), sheet_candidates)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [_safe_text(value) for value in rows[0]]
    return headers, rows[1:]


def _build_vector_text(document: dict[str, Any]) -> str:
    parts = [
        document["asset_category"],
        document["entity_name"],
        _display_text(document["entity_name"]),
        document["attribute_name"],
        _display_text(document["attribute_name"]),
        document["attribute_description"],
        document["entity_description"],
        document["source_description"],
        document["source_system"],
        document["asset_type"],
        document["type_classification"],
        document["business_classification"],
        document["category"],
        document["data_classification"],
        document["physical_region"],
    ]
    return " | ".join(part for part in parts if part)


def _build_semantic_text(document: dict[str, Any]) -> str:
    parts = [
        _display_text(document["attribute_name"]),
        document["attribute_description"],
        _display_text(document["entity_name"]),
        document["entity_description"],
    ]
    return " | ".join(part for part in parts if part)


def build_glossary_documents_from_workbook(workbook_path: Path | None = None) -> list[dict[str, Any]]:
    source_path = workbook_path or _resolve_glossary_workbook_path()
    headers, rows = _open_workbook_rows_from_path(source_path, GLOSSARY_SHEET_CANDIDATES)
    documents: list[dict[str, Any]] = []
    seen_doc_ids: dict[str, int] = {}

    for row in rows:
        row_map = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}

        layer = _safe_text(row_map.get("Layer")).upper()
        entity_name = _safe_text(row_map.get("business_term"))
        attribute_name = _safe_text(row_map.get("col_attributes"))
        attribute_description = _safe_text(row_map.get("Column Definition"))
        entity_description = _safe_text(row_map.get("Custom Table description"))
        source_description = _safe_text(row_map.get("Source Description"))
        source_system = _safe_text(row_map.get("Source System"))
        asset_type = _safe_text(row_map.get("Asset Type"))
        type_classification = _safe_text(row_map.get("Type Classification"))
        business_classification = _safe_text(row_map.get("Business Classification(Team)"))
        category = _safe_text(row_map.get("Category"))
        data_classification = _safe_text(row_map.get("Data Classification"))
        physical_region = _safe_text(row_map.get("Physical Region")).upper()
        is_pii_column = _safe_bool(row_map.get("Is PII Column"))
        is_sensitive_column = _safe_bool(row_map.get("Is Sensitive Column"))

        if not layer or not entity_name or not attribute_name:
            continue

        doc_id = f"attr::{layer.lower()}::{_normalize_name(entity_name)}::{_normalize_name(attribute_name)}"
        if physical_region:
            doc_id = f"{doc_id}::{_normalize_name(physical_region)}"

        seen_doc_ids[doc_id] = seen_doc_ids.get(doc_id, 0) + 1
        if seen_doc_ids[doc_id] > 1:
            doc_id = f"{doc_id}::{seen_doc_ids[doc_id]}"

        document: dict[str, Any] = {
            "doc_id": doc_id,
            "asset_category": layer,
            "asset_name": f"{layer}.{entity_name}",
            "asset_attribute": f"{entity_name}.{attribute_name}",
            "entity_name": entity_name,
            "attribute_name": attribute_name,
            "attribute_description": attribute_description,
            "entity_description": entity_description,
            "source_description": source_description,
            "source_system": source_system,
            "asset_type": asset_type,
            "type_classification": type_classification,
            "business_classification": business_classification,
            "category": category,
            "data_classification": data_classification,
            "physical_region": physical_region,
            "is_pii_column": is_pii_column,
            "is_sensitive_column": is_sensitive_column,
            "lineage_assets": _extract_lineage_assets(source_description),
            "join_keys": [attribute_name] if _is_join_key(attribute_name) else [],
            "vector_text": "",
            "semantic_text": "",
            "response_payload": {
                "asset_category": layer,
                "asset_name": f"{layer}.{entity_name}",
                "asset_attribute": f"{entity_name}.{attribute_name}",
            },
        }
        document["vector_text"] = _build_vector_text(document)
        document["semantic_text"] = _build_semantic_text(document)
        documents.append(document)

    return documents


def rebuild_glossary_json() -> Path:
    workbook_path = _resolve_glossary_workbook_path()
    output_path = _resolve_glossary_json_output_path()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    documents = build_glossary_documents_from_workbook(workbook_path)
    output_path.write_text(json.dumps(documents, indent=2, ensure_ascii=False), encoding="utf-8")
    return output_path


def ensure_glossary_json() -> Path:
    workbook_path = _resolve_glossary_workbook_path()
    output_path = _resolve_glossary_json_output_path()
    if not output_path.exists() or workbook_path.stat().st_mtime > output_path.stat().st_mtime:
        return rebuild_glossary_json()
    return output_path


def _normalize_requirement_header(text: str) -> str:
    return _normalize_name(text)


def _row_map_from_headers(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}


def _build_requirement_query(attribute_name: str, attribute_description: str, region: str) -> str:
    parts = [part for part in [_safe_text(attribute_name), _safe_text(attribute_description)] if part]
    if _safe_text(region):
        parts.append(f"Region: {_safe_text(region)}")
    return " | ".join(parts)


def _requirement_lookup(headers: list[str]) -> dict[str, str]:
    normalized_lookup = {_normalize_requirement_header(header): header for header in headers if header}
    lookup: dict[str, str] = {}
    for key, label in {**REQUIREMENT_REQUIRED_COLUMNS, **REQUIREMENT_OPTIONAL_COLUMNS}.items():
        header = normalized_lookup.get(_normalize_requirement_header(label))
        if header:
            lookup[key] = header
    missing = [label for key, label in REQUIREMENT_REQUIRED_COLUMNS.items() if key not in lookup]
    if missing:
        raise ValueError(
            "Requirement file is missing required column(s): "
            f"{', '.join(missing)}"
        )
    return lookup


def _parse_requirement_rows(headers: list[str], rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    lookup = _requirement_lookup(headers)
    requirements: list[dict[str, Any]] = []

    for row_index, row in enumerate(rows, start=2):
        row_map = _row_map_from_headers(headers, row)
        business_attribute = _safe_text(row_map.get(lookup["business_attribute"]))
        business_definition = _safe_text(row_map.get(lookup["business_definition"]))

        if not business_attribute and not business_definition:
            continue
        if not business_attribute or not business_definition:
            raise ValueError(
                f"Requirement row {row_index} must include both Business Attribute and Business Definition."
            )

        region = _safe_text(row_map.get(lookup.get("region", ""), "")) if lookup.get("region") else ""
        value_stream = _safe_text(row_map.get(lookup.get("value_stream", ""), "")) if lookup.get("value_stream") else ""
        brief_problem_statement = _safe_text(row_map.get(lookup.get("brief_problem_statement", ""), "")) if lookup.get("brief_problem_statement") else ""
        system_requirements = _safe_text(row_map.get(lookup.get("system_requirements", ""), "")) if lookup.get("system_requirements") else ""

        requirements.append(
            {
                "row_number": row_index,
                "business_attribute": business_attribute,
                "business_definition": business_definition,
                "region": region,
                "value_stream": value_stream,
                "brief_problem_statement": brief_problem_statement,
                "system_requirements": system_requirements,
                "attribute_name": business_attribute,
                "attribute_description": business_definition,
                "selected_value_stream": value_stream,
                "query": _build_requirement_query(business_attribute, business_definition, region),
            }
        )

    return requirements


def parse_requirement_workbook(file_bytes: bytes, file_name: str) -> list[dict[str, Any]]:
    lower_name = _safe_text(file_name).lower()
    if lower_name.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        headers = list(reader.fieldnames or [])
        rows = [tuple(record.get(header) for header in headers) for record in reader]
        return _parse_requirement_rows(headers, rows)

    headers, rows = _open_workbook_rows_from_bytes(file_bytes, REQUIREMENT_SHEET_CANDIDATES)
    return _parse_requirement_rows(headers, rows)
