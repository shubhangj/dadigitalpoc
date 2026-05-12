from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parent
SOURCE_WORKBOOK = Path(r"C:\Users\pilla\Desktop\HSBC_DATA_POC\DATA\Loan_MDA_GDA_CDA_Business_Glossary_Final_Product.xlsx")
LOCAL_WORKBOOK = REPO_ROOT / "DATA" / "Loan_MDA_GDA_CDA_Business_Glossary_Final_Product.xlsx"
OUTPUT_JSON = REPO_ROOT / "DATA" / "Business_Glossary_Output.json"
SHEET_NAME = "Glossary_All"


def _safe_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value != value:
        return ""
    return str(value).strip()


def _safe_bool(value):
    if isinstance(value, bool):
        return value
    text = _safe_text(value).lower()
    if text in {"true", "yes", "y", "1"}:
        return True
    if text in {"false", "no", "n", "0"}:
        return False
    return None


def _normalize_key(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def _display_text(text: str) -> str:
    return text.replace("_", " ")


ABBREVIATION_MAP = {
    "acct": "account",
    "addr": "address",
    "amt": "amount",
    "bal": "balance",
    "cd": "code",
    "coll": "collateral",
    "cust": "customer",
    "desc": "description",
    "dt": "date",
    "flg": "flag",
    "ind": "indicator",
    "id": "identifier",
    "ln": "loan",
    "nm": "name",
    "no": "number",
    "num": "number",
    "pct": "percent",
    "prov": "provision",
    "qty": "quantity",
    "ref": "reference",
    "seg": "segment",
    "stat": "status",
    "sts": "status",
    "txn": "transaction",
    "typ": "type",
}


def _expand_attribute_name(text: str) -> str:
    expanded_tokens = []
    for token in _normalize_key(text).split("_"):
        if not token:
            continue
        expanded_tokens.extend(ABBREVIATION_MAP.get(token, token).split("_"))
    return "_".join(expanded_tokens)


def _extract_lineage_assets(text: str) -> list[str]:
    assets: list[str] = []
    for token in re.findall(r"\b[A-Z][A-Z0-9_]{2,}\b", text or ""):
        if token in {"MDA", "GDA", "CDA", "UK"} or token in assets:
            continue
        assets.append(token)
    return assets


def _is_join_key(attribute_name: str) -> bool:
    key = _normalize_key(attribute_name)
    return key.endswith("_id") or key.endswith("_key")


def _resolve_workbook() -> Path:
    for path in (SOURCE_WORKBOOK, LOCAL_WORKBOOK):
        if path.exists():
            return path
    raise FileNotFoundError(
        "Loan analytics workbook not found. Expected one of: "
        f"{SOURCE_WORKBOOK} or {LOCAL_WORKBOOK}"
    )


def _build_vector_text(document: dict) -> str:
    parts = [
        document["asset_category"],
        document["entity_name"],
        _display_text(document["entity_name"]),
        document["attribute_name"],
        _display_text(document["attribute_name"]),
        document["attribute_description"],
        document["entity_description"],
        document["source_description"],
        document["source_system"],
        document["asset_type"],
        document["type_classification"],
        document["business_classification"],
        document["category"],
        document["data_classification"],
        document["physical_region"],
    ]
    return " | ".join(part for part in parts if part)


def _build_semantic_text(document: dict) -> str:
    parts = [
        _display_text(document["attribute_name"]),
        document["attribute_description"],
        _display_text(document["entity_name"]),
        document["entity_description"],
    ]
    return " | ".join(part for part in parts if part)


def build_documents() -> list[dict]:
    workbook_path = _resolve_workbook()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    row_iter = sheet.iter_rows(values_only=True)
    headers = [_safe_text(value) for value in next(row_iter, [])]
    documents: list[dict] = []
    seen_doc_ids: dict[str, int] = {}

    for row in row_iter:
        row_map = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}

        layer = _safe_text(row_map.get("Layer")).upper()
        entity_name = _safe_text(row_map.get("business_term"))
        attribute_name = _safe_text(row_map.get("col_attributes"))
        attribute_description = _safe_text(row_map.get("Column Definition"))
        entity_description = _safe_text(row_map.get("Custom Table description"))
        source_description = _safe_text(row_map.get("Source Description"))
        source_system = _safe_text(row_map.get("Source System"))
        asset_type = _safe_text(row_map.get("Asset Type"))
        type_classification = _safe_text(row_map.get("Type Classification"))
        business_classification = _safe_text(row_map.get("Business Classification(Team)"))
        category = _safe_text(row_map.get("Category"))
        data_classification = _safe_text(row_map.get("Data Classification"))
        physical_region = _safe_text(row_map.get("Physical Region")).upper()
        is_pii_column = _safe_bool(row_map.get("Is PII Column"))
        is_sensitive_column = _safe_bool(row_map.get("Is Sensitive Column"))

        if not layer or not entity_name or not attribute_name:
            continue

        doc_id = (
            f"attr::{layer.lower()}::{_normalize_key(entity_name)}::{_normalize_key(attribute_name)}"
        )
        if physical_region:
            doc_id = f"{doc_id}::{_normalize_key(physical_region)}"

        seen_doc_ids[doc_id] = seen_doc_ids.get(doc_id, 0) + 1
        if seen_doc_ids[doc_id] > 1:
            doc_id = f"{doc_id}::{seen_doc_ids[doc_id]}"

        document = {
            "doc_id": doc_id,
            "asset_category": layer,
            "asset_name": f"{layer}.{entity_name}",
            "asset_attribute": f"{entity_name}.{attribute_name}",
            "entity_name": entity_name,
            "attribute_name": attribute_name,
            "normalized_attribute_name": _normalize_key(attribute_name),
            "expanded_attribute_name": _expand_attribute_name(attribute_name),
            "attribute_description": attribute_description,
            "entity_description": entity_description,
            "source_description": source_description,
            "source_system": source_system,
            "asset_type": asset_type,
            "type_classification": type_classification,
            "business_classification": business_classification,
            "category": category,
            "data_classification": data_classification,
            "physical_region": physical_region,
            "is_pii_column": is_pii_column,
            "is_sensitive_column": is_sensitive_column,
            "lineage_assets": _extract_lineage_assets(source_description),
            "join_keys": [attribute_name] if _is_join_key(attribute_name) else [],
            "vector_text": "",
            "semantic_text": "",
            "response_payload": {
                "asset_category": layer,
                "asset_name": f"{layer}.{entity_name}",
                "asset_attribute": f"{entity_name}.{attribute_name}",
            },
        }
        document["vector_text"] = _build_vector_text(document)
        document["semantic_text"] = _build_semantic_text(document)
        documents.append(document)

    return documents


def main() -> None:
    workbook_path = _resolve_workbook()
    documents = build_documents()
    OUTPUT_JSON.write_text(json.dumps(documents, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Source workbook: {workbook_path}")
    print(f"Wrote {len(documents)} analytics documents to {OUTPUT_JSON}")
    if documents:
        print(json.dumps(documents[0], indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
